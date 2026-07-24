"""Generate styled Excel (.xlsx) workbooks for applications export."""
from __future__ import annotations

import io
from typing import Iterable

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from ..db import Application

_STATUS_RU = {
    "pending": "Kutilmoqda / На рассмотрении",
    "approved": "Tasdiqlangan / Одобрено",
    "rejected": "Rad etilgan / Отклонено",
}


def generate_excel(apps: Iterable[Application]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Arizalar"

    # Header styling
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4C1D95", end_color="4C1D95", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center")

    headers = [
        "ID",
        "Reg. №",
        "Status",
        "F.I.SH / Имя",
        "Telegram Username",
        "Telegram ID",
        "Telefon / Телефон",
        "Davlat raqami / Гос. номер",
        "Yo'nalish / Направление",
        "Mamlakat / Страна",
        "Til / Язык",
        "Topshirilgan vaqt / Подана",
        "Ko'rib chiqilgan vaqt / Обработана",
        "Moderator",
    ]

    ws.append(headers)
    ws.row_dimensions[1].height = 28

    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = thin_border

    # Fills for status
    status_fills = {
        "pending": PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"),
        "approved": PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid"),
        "rejected": PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"),
    }
    status_fonts = {
        "pending": Font(name="Calibri", size=10, bold=True, color="92400E"),
        "approved": Font(name="Calibri", size=10, bold=True, color="065F46"),
        "rejected": Font(name="Calibri", size=10, bold=True, color="991B1B"),
    }

    body_font = Font(name="Calibri", size=10)

    for r_idx, app in enumerate(apps, start=2):
        row_data = [
            app.id,
            f"№{app.reg_number}" if app.reg_number is not None else "—",
            _STATUS_RU.get(app.status, app.status),
            app.full_name or "—",
            f"@{app.username.lstrip('@')}" if app.username else "—",
            app.user_id,
            app.phone or "—",
            app.plate or "—",
            app.direction or "—",
            app.country or "—",
            app.language.upper(),
            app.created_at,
            app.processed_at or "—",
            app.processed_by or "—",
        ]
        ws.append(row_data)
        ws.row_dimensions[r_idx].height = 22

        for c_idx in range(1, len(row_data) + 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.font = body_font
            cell.border = thin_border
            cell.alignment = align_left if c_idx not in (1, 2, 3, 6, 11) else align_center

            # Highlight status column
            if c_idx == 3 and app.status in status_fills:
                cell.fill = status_fills[app.status]
                cell.font = status_fonts[app.status]

    # Auto-fit columns
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or "")
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
